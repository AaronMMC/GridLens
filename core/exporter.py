"""
CSV and Excel export utilities for extracted table data.

Call ``export_csv()`` or ``export_excel()`` with the same ``data`` dict
structure returned by ``core.extractor.extract_table()``:

.. code:: python

    {"headers": ["Col1", "Col2"], "rows": [["a", "b"], ["c", "d"]]}
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def export_csv(data: dict, path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(data["headers"])
        w.writerows(data["rows"])


def export_excel(data: dict, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    fill = PatternFill("solid", fgColor="BDD7EE")
    for ci, h in enumerate(data["headers"], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = fill
    for ri, row in enumerate(data["rows"], 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 50)
    wb.save(path)